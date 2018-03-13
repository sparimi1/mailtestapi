'''
Created on Mar 12, 2018

@author: Administrator
'''

import openpyxl 

def main():
    wb = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\API testing\\MailTest.xlsx')
    
    print( wb.get_sheet_names())
    # ['Sheet2', 'New Title', 'Sheet1']
    # Get a sheet to read
    sheet = wb.worksheets[0]
    # No of written Rows in sheet
    r = sheet.max_row
    print('this is max rows' + str(r))
    # No of written Columns in sheet
    c = sheet.max_column
    print('max column size'+ str(c))
    # Reading each cell in excel
    for i in range(1, r+1):
        for j in range(1, c+1):
            print ('the value of cell' + str(i),str(j) + 'is' + str(sheet.cell(row=i, column=j).value))
            

if __name__ == '__main__':
    main()