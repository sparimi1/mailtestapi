'''
Created on Mar 11, 2018

@author: Administrator
'''
import json

import requests 

import urllib

import openpyxl


main_API  =  'https://api.mailtest.in/v1/'

domain = 'active.mailtest.in'

URL = main_API + domain



wb = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\API testing\\MailTest.xlsx')

sheet = wb.get_sheet_by_name('MailTest')

print(sheet)

print(sheet.title)

print(sheet.max_row)

print(sheet.max_column)
 
ws = wb['MailTest']

for rowNum in range(2,sheet.max_row+1,1):
    
        URL = ( (sheet.cell(row = rowNum , column = 1).value) + (sheet.cell(row = rowNum , column = 2).value))
        print(URL)
        
        json_data = requests.get(URL).json()
        print(json_data)
        json_status = json_data['status']
        json_message = json_data["message"]
        json_code  = json_data['code']
        # Write the json response into different cells of excel sheet 
        (sheet.cell(row = rowNum , column = 3).value) = json_code
        (sheet.cell(row = rowNum , column = 4).value) = json_message
        (sheet.cell(row = rowNum , column = 5).value) = json_status
        
        # Test criteria check 
        if (
        (sheet.cell(row = rowNum , column = 3).value) == (sheet.cell(row = rowNum , column = 6).value)):
          (sheet.cell(row = rowNum , column = 9).value) = 'PASSED'
        else:
            (sheet.cell(row = rowNum , column = 9).value) = 'FAILED'

        print('The status of domain' + (sheet.cell(row =rowNum , column = 2).value) + '     is    '  + '    '  + json_status)
        
# Save the test run results into a new sheet so we do not loose the original input sheet.                 
wb.save('C:\\Users\\Administrator\\Desktop\\API testing\\UpdatedMailTest.xlsx')
     
      
        
        
        
        
        
        
        